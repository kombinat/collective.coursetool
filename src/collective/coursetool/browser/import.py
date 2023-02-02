from collective.coursetool import _
from collective.coursetool.config import BASE_FOLDER_ID
from collective.coursetool.content.member import IMemberSchema
from collective.coursetool.interfaces import IImportingMembers
from collective.coursetool.utils import generate_member_id
from datetime import datetime
from openpyxl import load_workbook
from plone import api
from plone.restapi.interfaces import IDeserializeFromJson
from plone.schema.email import _isemail
from zope.annotation.interfaces import IAnnotations
from zope.component import getMultiAdapter
from zope.interface import alsoProvides
from zope.interface import Interface
from zope.publisher.browser import BrowserView

import logging
import transaction


log = logging.getLogger(__name__)


SCHEMA_MAPPING = {
    # schema name: excel col idx
    "customer_id": 1,
    "salutation": 2,
    "graduation": 3,
    "last_name": 4,
    "first_name": 5,
    "address": 7,
    "address2": 6,
    "cty_code": 8,
    "zip_code": 9,
    "city": 10,
    "website": 11,
    "booking_nr": 12,
    "inactive": 13,
    "email": 14,
    "phone": 15,
    "mobile_phone": 16,
    "fax": 17,
    "birthday": 18,
    "salutation_letter": 21,
    "pass_issue_date": 23,
    "pass_expiration_date": 24,
    "payed": 25,
    "payed_date": 26,
    "salutation_personal": 27,
}
VOCAB_MAPPING = {
    "state": 31,
    "qualification": 32,
    "partner_type": 33,
}
METADATA_MAPPING = {
    "created": 19,
    "modified": 20,
    "effective": 23,
    "expired": 24,
}
STATE_MAPPING = {
    "Eingelesen mit Befähigung": "read_qualified",
    "Eingelesen, KEINE Befähigung": "read_no_qualification",
    "Eingelesen, KEIN Geburtsdatum": "read_no_birthday",
    "Komplett, NICHT gedruckt": "complete_not_printed",
    "Komplett, gedruckt": "complete_printed",
    "Nicht Eingelesen": "not_read",
}
SALUTATION_MAPPING = {
    "Herrn": "male",
    "Herr": "male",
    "Frau": "female",
}
DISABLE_VALIDATION = [
    "email", "birthday", "first_name", "last_name"
]

def no_validation(value):
    return


class ImportMembers(BrowserView):
    def __call__(self):
        if not self.request.get("import_data"):
            return self.index()

        alsoProvides(self.request, IImportingMembers)

        _xlsx_file = self.request["import_data"]

        wb = load_workbook(_xlsx_file, read_only=True)
        ws = wb.active
        ws.reset_dimensions()

        schema_names = IMemberSchema.names()
        member_base = api.portal.get()[BASE_FOLDER_ID]["members"]
        customer_id_map = {
            b.customer_id: b.id
            for b in self.context.portal_catalog(portal_type="coursetool.member")
        }

        # disable constraints for vocabularies during import
        for fld in VOCAB_MAPPING.keys():
            schema_fld = IMemberSchema[fld]
            if schema_fld._type == tuple:
                schema_fld.value_type._orig_validate = schema_fld.value_type._validate
                schema_fld.value_type._validate = no_validation
            else:
                schema_fld._orig_validate = schema_fld._validate
                schema_fld._validate = no_validation

        # disable required fields during import
        for fld in DISABLE_VALIDATION:
            IMemberSchema[fld]._orig_required = IMemberSchema[fld].required
            IMemberSchema[fld]._orig_validate = IMemberSchema[fld]._validate
            IMemberSchema[fld].required = False
            IMemberSchema[fld]._validate = no_validation

        # assume the first row columns are titles
        for num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            data = {}

            for fld, idx in SCHEMA_MAPPING.items():
                if fld not in schema_names:
                    continue
                if IMemberSchema[fld]._type == bool:
                    data[fld] = str(row[idx]).lower() == "true"
                elif IMemberSchema[fld]._type == str:
                    # force str
                    data[fld] = str(row[idx]).strip() if row[idx] else ""
                elif isinstance(row[idx], datetime):
                    data[fld] = row[idx].strftime("%Y-%m-%d") if row[idx] else None
                else:
                    data[fld] = row[idx]

                # extra mappings
                if fld == "salutation":
                    data[fld] = SALUTATION_MAPPING.get(row[idx], None) if row[idx] else None

                if (
                    fld == "email"
                    and data[fld] != ""
                ):
                    if not _isemail(data[fld]):
                        # handle broken mail addresses
                        log.warn("EMail not valid: %s", row[idx])
                        data[fld] = None
                    else:
                        data[fld] = data[fld].lower()

            for fld, idx in METADATA_MAPPING.items():
                data[fld] = row[idx].isoformat() if row[idx] else None

            for fld, idx in VOCAB_MAPPING.items():
                if len(row) <= idx :
                    continue
                if fld == "state":
                    data[fld] = STATE_MAPPING.get(row[idx], None) if row[idx] else None
                    continue
                data[fld] = [v.strip() for v in row[idx].split(",") if v] if row[idx] else None

            __traceback_info__ = data

            if data["customer_id"] in customer_id_map:
                obj = member_base[customer_id_map[data["customer_id"]]]
                msg = "{0} Imported data for {1}"
            else:
                obj = api.content.create(
                    container=member_base,
                    type="coursetool.member",
                    id=generate_member_id()
                )
                msg = "{0} Generated new member {1}"

            deserializer = getMultiAdapter((obj, self.request), IDeserializeFromJson)
            try:
                obj = deserializer(validate_all=False, data=data)
                obj.reindexObject()
                log.info(msg.format(num, obj))
            except Exception as msg:
                log.warn(f"Could not set data {data} for {obj}: {msg}")

            transaction.commit()

        # reset original constraints
        for fld in VOCAB_MAPPING.keys():
            schema_fld = IMemberSchema[fld]
            if schema_fld._type == tuple:
                schema_fld.value_type._validate = schema_fld.value_type._orig_validate
                delattr(schema_fld.value_type, "_orig_validate")
            else:
                schema_fld._validate = schema_fld._orig_validate
                delattr(schema_fld, "_orig_validate")

        # reset required fields
        for fld in DISABLE_VALIDATION:
            IMemberSchema[fld].required = IMemberSchema[fld]._orig_required
            IMemberSchema[fld]._validate = IMemberSchema[fld]._orig_validate
            delattr(IMemberSchema[fld], "_orig_required")
            delattr(IMemberSchema[fld], "_orig_validate")

        api.portal.show_message(_("Imported members"), request=self.request)
        return self.index()
